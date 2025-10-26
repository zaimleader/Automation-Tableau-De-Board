import json
import sys
from openpyxl import load_workbook

# --- Read column name from argument ---
if len(sys.argv) < 2:
    print("Error: Column name not provided.")
    sys.exit(1)

column_letter = sys.argv[1].upper()

# --- Load Gmail data ---
with open(r"C:\Users\pc\Desktop\Documents\Programming\Automation Excel\main\result\result.json", "r", encoding="utf-8") as f:
    data = json.load(f)

cal_ref = extracted.get("cal_ref")
sw_ref = extracted.get("sw_ref")
hw_ref = extracted.get("hw_ref")

# --- Open the Excel file ---
excel_path = r"C:\Users\pc\Desktop\Documents\Programming\Automation Excel\excel\TDB_CP_BPILOT_V4 (12).xlsm"
wb = load_workbook(excel_path)
ws = wb.active

# --- Search for the last occurrence of the name ---
last_row = None
for row in range(1, ws.max_row + 1):
    cell_value = str(ws[f"{column_letter}{row}"].value)
    if name.lower() in cell_value.lower():
        last_row = row

if last_row:
    cell = ws[f"{column_letter}{last_row}"]
    cell.value = f"     {cell.value} {status}"  # add 5 spaces before name + DONE
    print(f"Updated row {last_row}: {cell.value}")
else:
    print(f"Name '{name}' not found in column {column_letter}.")

# --- Save the updated file ---
wb.save(excel_path)
print("Excel file updated successfully.")
